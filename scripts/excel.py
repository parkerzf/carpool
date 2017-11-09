from __future__ import print_function
import os, sys
import numpy as np
import pandas as pd
from openpyxl import load_workbook

# python scripts/excel.py carpool.log AppMatch threshold_R51_0.3 40 2 50
log_path = sys.argv[1]
app = sys.argv[2]
graph_base = sys.argv[3]
num_user = int(sys.argv[4])
num_spot = int(sys.argv[5])
num_node = int(sys.argv[6])

basedir = os.path.dirname(os.path.realpath(__file__))
dir_path = '{}/../data/{}_{}_{}_{}'.format(basedir, graph_base, num_user, num_spot, num_node)

dic = {'AppBasic': 23, 'AppStaticMatch': 24, 'AppMatch': 25, 'AppStaticTaxiMatch': 26}

def update_excel(row):
	excel_file_base_path = '{}.xlsm'.format(os.path.splitext(row[2].strip())[0])
	excel_file_path = '{}/{}'.format(dir_path, excel_file_base_path)

	cost = row[4]
	wb = load_workbook(excel_file_path, keep_vba=True)
	ws = wb['AIMMS Output']
	ws['A{}'.format(dic[app])] = app
	ws['B{}'.format(dic[app])] = float(cost)
	wb.save(excel_file_path)

df = pd.read_csv(log_path, sep='|', header=None)

df_filtered = df[(df[1].str.strip() == app) & (df[2].str.strip().str.startswith('threshold'))]

df_filtered.apply(update_excel, axis=1)
